import networkx as nx
import openpyxl
import xlsxwriter
import numpy as np
import os

file = "updated_original.xlsx"

_book = openpyxl.load_workbook(file)
_sheet = _book.active

allOutgoingNodes = list()
noRepeatitionNodes = list()

for row in _sheet.iter_rows(min_row=2, values_only=True):
    allOutgoingNodes.append(row[0])

noRepeatitionNodes = list(dict.fromkeys(allOutgoingNodes))

print(noRepeatitionNodes)
flag = 1
pr = None

output_dir = "simulation_output_collectively"
os.makedirs(output_dir, exist_ok=True)

#Category order: Education, Family, Core, Work, Skills, Social, relationships
nodes_to_observe_their_effects = ['value of wellbeing recongized in school culture', 'time to spend with family', 'self-awareness', 'time and energy for non-work activity', 'strength of friendships', 'socio-economic opportunities', 'resilience']


combined_output_file = os.path.join(output_dir, f"combined_effect_{file}")
workbook_combined = xlsxwriter.Workbook(combined_output_file, {'nan_inf_to_errors': True})
worksheet_combined = workbook_combined.add_worksheet()

j = 1

for w in np.arange(1, 100, 1):
    print(w)
    G = nx.DiGraph()
    weight = 0
    for row in _sheet.iter_rows(min_row=2, values_only=True):
        _Node1 = row[0]
        _Node2 = row[1]
        if _Node2 in nodes_to_observe_their_effects:
            weight = float(row[4]) * float(w)
        else:
            weight = float(row[4])
        G.add_weighted_edges_from([(_Node1, _Node2, weight)])

    nodes_list = list(G.nodes())
    pr = nx.pagerank_numpy(G, alpha=0.5)
    pr_keys = list(pr.keys())
    pr_values = list(pr.values())
    my_formatted_list = ['%.4f' % elem for elem in pr_values]
    if w == 1:
        for col_num, data in enumerate(pr_keys):
            worksheet_combined.write(0, col_num, data)
    for col_num, data in enumerate(my_formatted_list):
        try:
            worksheet_combined.write(j, col_num, float(data))
        except ValueError:
            worksheet_combined.write(j, col_num, 'NaN')
    j += 1
    G.clear()
    del G

workbook_combined.close()
