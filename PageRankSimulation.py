import networkx as nx
import openpyxl
import xlsxwriter
import numpy as np
import os

file = "updated_original.xlsx"

_book = openpyxl.load_workbook(file)
_sheet = _book.active

allOutgoingNodes = list()
noRepeatitionNodes = list()

for row in _sheet.iter_rows(min_row=2, values_only=True):
    allOutgoingNodes.append(row[0])

noRepeatitionNodes = list(dict.fromkeys(allOutgoingNodes))

print(noRepeatitionNodes)
flag = 1
pr = None

# Create directory if it doesn't exist
output_dir = "simulation_Output"
os.makedirs(output_dir, exist_ok=True)

for _nodes in noRepeatitionNodes:
    j = 1
    occurrences = allOutgoingNodes.count(_nodes)
    workbook = xlsxwriter.Workbook(os.path.join(output_dir, f"{_nodes}_{file}"), {'nan_inf_to_errors': True})
    worksheet = workbook.add_worksheet()

    for w in np.arange(1, 100, 1):
        print(w)
        G = nx.DiGraph()
        weight = 0
        for row in _sheet.iter_rows(min_row=2, values_only=True):
            _Node1 = row[0]
            _Node2 = row[1]
            if _Node2 == _nodes:
                weight = float(row[3]) * float(w)
            elif _Node2 != _nodes:
                weight = float(row[3])
            G.add_weighted_edges_from([(_Node1, _Node2, weight)])

        nodes_list = list(G.nodes())
        pr = nx.pagerank_numpy(G, alpha=0.65)
        pr_keys = list(pr.keys())
        pr_values = list(pr.values())
        my_formatted_list = ['%.4f' % elem for elem in pr_values]
        if w == 1:
            for col_num, data in enumerate(pr_keys):
                worksheet.write(0, col_num, data)
        for col_num, data in enumerate(my_formatted_list):
            try:
                worksheet.write(j, col_num, float(data))
            except ValueError:
                worksheet.write(j, col_num, 'NaN')
        j += 1
        G.clear()
        del G
    workbook.close()
